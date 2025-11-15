# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook

# Пути к файлам
GENERATED_XLSX = "generated.xlsx"
PROCOUNTER_XLSX = "procounter.xlsx"

# Читаем весь procounter.xlsx (берём первый лист)
df_pro = pd.read_excel(PROCOUNTER_XLSX)

# Загружаем основной файл
wb = load_workbook(GENERATED_XLSX)

# Если лист с таким именем уже существует — удалим
if "procounter" in wb.sheetnames:
    ws_old = wb["procounter"]
    wb.remove(ws_old)

# Создаем новый лист
ws = wb.create_sheet("procounter")

# Записываем DataFrame в созданный лист
for r_idx, row in enumerate(df_pro.itertuples(index=False), start=2):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Записываем заголовки
for col_idx, col_name in enumerate(df_pro.columns, start=1):
    ws.cell(row=1, column=col_idx, value=col_name)

wb.save(GENERATED_XLSX)

print("✅ Лист 'procounter' успешно добавлен в generated.xlsx")
