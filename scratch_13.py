# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

GENERATED_XLSX = "generated.xlsx"

wb = load_workbook(GENERATED_XLSX)
ws = wb.active

start_row = 8
row = start_row

# Колонки, которым ставим формат 0.00 для КАЖДОЙ строки с формулами
formula_columns = ["E","F","G","H","I","J","K","L","M","N",
                   "AN","AU","AR","AV","AX","AY","AZ","BA","BB","BC",
                   "BD","BE","BG","BH","BI","BJ","BK","BL","BM"]

while True:
    cell_A = ws[f"A{row}"].value
    if cell_A is None or str(cell_A).strip() == "":
        break  # конец данных

    # 1) E = сумма F:N
    ws[f"E{row}"] = f"=SUM(F{row}:N{row})"

    # 2) F = P + Z
    ws[f"F{row}"] = f"=P{row}+Z{row}"

    # 3) G = Q + AA
    ws[f"G{row}"] = f"=Q{row}+AA{row}"

    # 4) H = V + AF
    ws[f"H{row}"] = f"=V{row}+AF{row}"

    # 5) I = W + AG
    ws[f"I{row}"] = f"=W{row}+AG{row}"

    # 6) J = R + AB
    ws[f"J{row}"] = f"=R{row}+AB{row}"

    # 7) K = S + AC
    ws[f"K{row}"] = f"=S{row}+AC{row}"

    # 8) L = X + AH
    ws[f"L{row}"] = f"=X{row}+AH{row}"

    # 9) M = T + AD
    ws[f"M{row}"] = f"=T{row}+AD{row}"

    # 10) N = O + Y
    ws[f"N{row}"] = f"=O{row}+Y{row}"

    # 11) AN = AN7 * AO(row)
    ws[f"AN{row}"] = f"=$AN$7*AO{row}"

    # 12) AU = D
    ws[f"AU{row}"] = f"=D{row}"

    # 13) AR = D * AQ
    ws[f"AR{row}"] = f"=D{row}*AQ{row}"

    # 14) AV = (AU - AU*AW - AU*$AX$5 - AU*$AY$5)
    ws[f"AV{row}"] = f"=(AU{row}-(AU{row}*AW{row})-(AU{row}*$AX$5)-(AU{row}*$AY$5))"

    # 15) AX = большая формула
    ws[f"AX{row}"] = (
        f"=(E{row}*D{row})"
        f"+(F{row}*D{row}/2)"
        f"+(G{row}*D{row})"
        f"+((H{row}*D{row})*1.5)"
        f"+((I{row}*D{row})*2)"
        f"+((J{row}*D{row})*3)"
        f"+(K{row}*$K$7)"
        f"+(L{row}*$L$7)"
        f"+(M{row}*D{row})"
        f"+AR{row}"
        f"+AS{row}"
    )

    # 16) AY = (AX - AR) * $AW$5
    ws[f"AY{row}"] = f"=(AX{row}-AR{row})*$AW$5"

    # 17) AZ = AX * $AV$5
    ws[f"AZ{row}"] = f"=AX{row}*$AV$5"

    # 18) BA = ((AX+AY+AZ)*AW)*(-1)
    ws[f"BA{row}"] = f"=((AX{row}+AY{row}+AZ{row})*(AW{row}))*(-1)"

    # 19) BB = ((AX+AY+AZ)*$BA$5)*(-1)  <- твоя поправка
    ws[f"BB{row}"] = f"=((AX{row}+AY{row}+AZ{row})*$BA$5)*(-1)"

    # 20) BC = ((AX+AY+AZ)*$AX$5)*(-1)
    ws[f"BC{row}"] = f"=((AX{row}+AY{row}+AZ{row})*$AX$5)*(-1)"

    # 21) BD = (((AX+AY+AZ)) - (BA*(-1)) - (BB*(-1)) - (BC*(-1)) - (AJ*(-1)) + AN) - AT
    ws[f"BD{row}"] = (
        f"=(((AX{row}+AY{row}+AZ{row}))"
        f"-((BA{row})*(-1))"
        f"-((BB{row})*(-1))"
        f"-((BC{row})*(-1))"
        f"-((AJ{row})*(-1))"
        f"+AN{row})"
        f"-AT{row}"
    )

    # 22) BE = AX + AY + AZ
    ws[f"BE{row}"] = f"=AX{row}+AY{row}+AZ{row}"

    # 23) BG = BE * $BG$5
    ws[f"BG{row}"] = f"=BE{row}*$BG$5"

    # 24) BH = BE * $BH$5
    ws[f"BH{row}"] = f"=BE{row}*$BH$5"

    # 25) BI = BE * $BI$5
    ws[f"BI{row}"] = f"=BE{row}*$BI$5"

    # 26) BJ = BE * $BJ$5
    ws[f"BJ{row}"] = f"=BE{row}*$BJ$5"

    # 27) BK = BE * $BK$5
    ws[f"BK{row}"] = f"=BE{row}*$BK$5"

    # 28) BL = BG+BH+BI+BJ+BK
    ws[f"BL{row}"] = f"=BG{row}+BH{row}+BI{row}+BJ{row}+BK{row}"

    # 29) BM = BG+BH+BI+BJ+BK+BD
    ws[f"BM{row}"] = f"=BG{row}+BH{row}+BI{row}+BJ{row}+BK{row}+BD{row}"

    # формат 0.00 для этих колонок
    for col in formula_columns:
        ws[f"{col}{row}"].number_format = "0.00"

    row += 1

# --------- ИТОГОВАЯ СТРОКА TOTAL ---------
total_row = row             # сюда ставим итоги
first_row = start_row
last_row = row - 1

# подпись TOTAL (можно в A или D — как хочешь)
ws[f"D{total_row}"] = "TOTAL"
ws[f"D{total_row}"].font = Font(bold=True)

# стили для зелёной строки
thin = Side(style="thin")
thick = Side(style="medium")
border_all = Border(top=thick, bottom=thick, left=thick, right=thick)
fill_green = PatternFill(fill_type="solid", fgColor="CCFFCC")  # бледно-зелёный

# проходим по всем колонкам от E до BM включительно
start_col_idx = column_index_from_string("E")
end_col_idx = column_index_from_string("BM")

for col_idx in range(start_col_idx, end_col_idx + 1):
    col_letter = get_column_letter(col_idx)
    cell_ref = f"{col_letter}{total_row}"
    cell = ws[cell_ref]

    cell.value = f"=SUM({col_letter}{first_row}:{col_letter}{last_row})"
    cell.number_format = "0.00"
    cell.font = Font(bold=True)
    cell.border = border_all
    cell.fill = fill_green

wb.save(GENERATED_XLSX)
print(f"✅ Формулы проставлены до строки {last_row}, TOTAL в строке {total_row} (E:BM).")
