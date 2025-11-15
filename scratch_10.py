# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook

# ----- ПУТИ -----
WEEKS_XLSX = r"tyontekijat_weeks.xlsx"   # файл с 1st week / 2nd week
GENERATED_XLSX = r"generated.xlsx"       # файл, куда писать людей, начиная с 8-й строки

# ===== ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ ПОИСКА КОЛОНКИ =====
def resolve_col(df_columns, logical_name: str):
    """
    Находит 'реальное' имя столбца в DataFrame для логического имени.
    Работает и для '50%'/'100%' (где в df колонка может называться 0.5 или 1).
    """
    cols = list(df_columns)

    # Логика для процентов типа '50%', '100%', '150%', ...
    if logical_name.endswith("%"):
        try:
            percent_val = float(logical_name.rstrip("%")) / 100.0  # 50% -> 0.5, 100% -> 1.0
        except ValueError:
            percent_val = None

        # 1) точное совпадение по строке
        for c in cols:
            if str(c).strip() == logical_name:
                return c

        if percent_val is not None:
            # 2) точное совпадение по числовому имени
            for c in cols:
                if isinstance(c, (int, float)) and abs(float(c) - percent_val) < 1e-9:
                    return c

            # 3) строковое представление числа '0.5', '1', '1.5' и т.п.
            for c in cols:
                try:
                    if abs(float(str(c).replace(",", ".")) - percent_val) < 1e-9:
                        return c
                except Exception:
                    continue

        # 4) fallback — поиск по вхождению '50%' в имени колонки (на всякий случай)
        target = logical_name.replace(" ", "").lower()
        for c in cols:
            norm = str(c).replace(" ", "").lower()
            if target in norm:
                return c

        return None

    # Логика для текстовых колонок (Norm, Evening shift bonus и т.п.)
    target = logical_name.replace(" ", "").lower()

    # 1) точное совпадение
    for c in cols:
        norm = str(c).replace(" ", "").lower()
        if norm == target:
            return c

    # 2) частичное совпадение (например 'eveningshiftbonus' в 'Evening shift bonus (1)')
    for c in cols:
        norm = str(c).replace(" ", "").lower()
        if target in norm:
            return c

    return None


# ===== ЧИТАЕМ ОБА ЛИСТА =====
df_w1_full = pd.read_excel(WEEKS_XLSX, sheet_name="1st week")
df_w2_full = pd.read_excel(WEEKS_XLSX, sheet_name="2nd week")

# Берём только ключи для списка людей
df_w1_keys = df_w1_full[["Name", "Surname"]].copy()
df_w2_keys = df_w2_full[["Name", "Surname"]].copy()

# ----- ОБЪЕДИНЯЕМ ВСЕХ ЛЮДЕЙ БЕЗ ДУБЛИКАТОВ -----
keys_w1 = df_w1_keys.apply(tuple, axis=1)
mask_new = ~df_w2_keys.apply(tuple, axis=1).isin(keys_w1)
df_w2_unique = df_w2_keys[mask_new]

df_all = pd.concat([df_w1_keys, df_w2_unique], ignore_index=True)
df_all = df_all.drop_duplicates(subset=["Name", "Surname"], keep="first").reset_index(drop=True)

# ----- ГОТОВИМ ДОСТУП ПО КЛЮЧУ (Name, Surname) -----
df_w1_map = df_w1_full.set_index(["Name", "Surname"])
df_w2_map = df_w2_full.set_index(["Name", "Surname"])

# ===== ОТКРЫВАЕМ generated.xlsx =====
wb = load_workbook(GENERATED_XLSX)
ws = wb.active   # при необходимости здесь выбери конкретный лист: wb["Sheet1"]

start_row = 8       # с какой строки начинаем писать
col_num = 1         # A
col_name = 2        # B
col_surname = 3     # C

# ----- Маппинг логических полей в Excel-колонки -----
# 1) из 1st week -> O..X
col_map_w1 = {
    "Norm": 15,                  # O
    "50%": 16,                   # P
    "100%": 17,                  # Q
    "300%": 18,                  # R
    "Evening shift bonus": 19,   # S
    "Urakka": 20,                # T
    "Sick leaves": 21,           # U
    "150%": 22,                  # V
    "200%": 23,                  # W
    "Night shift bonus": 24      # X
}

# 2) из 2nd week -> Y..AH
col_map_w2 = {
    "Norm": 25,                  # Y
    "50%": 26,                   # Z
    "100%": 27,                  # AA
    "300%": 28,                  # AB
    "Evening shift bonus": 29,   # AC
    "Urakka": 30,                # AD
    "Sick leaves": 31,           # AE
    "150%": 32,                  # AF
    "200%": 33,                  # AG
    "Night shift bonus": 34      # AH
}

# ===== ПРОХОДИМ ПО СПИСКУ ЛЮДЕЙ И ЗАПОЛНЯЕМ СТРОКИ =====
for idx, row in df_all.iterrows():
    excel_row = start_row + idx
    name = row["Name"]
    surname = row["Surname"]
    key = (name, surname)

    # A: нумерация
    ws.cell(row=excel_row, column=col_num, value=idx + 1)
    # B, C: Name, Surname
    ws.cell(row=excel_row, column=col_name,   value=name)
    ws.cell(row=excel_row, column=col_surname, value=surname)

    # ---------- 1) ДАННЫЕ ИЗ 1st week (O..X) ----------
    if key in df_w1_map.index:
        src1 = df_w1_map.loc[key]
        # если вдруг несколько строк с одним и тем же именем -> возьмём первую
        if isinstance(src1, pd.DataFrame):
            src1 = src1.iloc[0]

        for logical_name, excel_col in col_map_w1.items():
            actual_col = resolve_col(df_w1_map.columns, logical_name)
            if actual_col is not None:
                ws.cell(row=excel_row, column=excel_col, value=src1[actual_col])

    # ---------- 2) ДАННЫЕ ИЗ 2nd week (Y..AH) ----------
    if key in df_w2_map.index:
        src2 = df_w2_map.loc[key]
        if isinstance(src2, pd.DataFrame):
            src2 = src2.iloc[0]

        for logical_name, excel_col in col_map_w2.items():
            actual_col = resolve_col(df_w2_map.columns, logical_name)
            if actual_col is not None:
                ws.cell(row=excel_row, column=excel_col, value=src2[actual_col])

# Сохраняем результат
wb.save(GENERATED_XLSX)
print(f"✅ Записано {len(df_all)} строк в {GENERATED_XLSX} начиная с A{start_row}")
